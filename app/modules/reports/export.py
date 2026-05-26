"""Geração de relatórios em PDF e Excel."""

from __future__ import annotations

import io
from datetime import date, datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from typing import Any


# ---------------------------------------------------------------------------
# Cores do tema (RGB)
# ---------------------------------------------------------------------------
_PRIMARY = (59, 130, 246)       # #3b82f6 – blue-500
_PRIMARY_DARK = (37, 99, 235)   # #2563eb – blue-600
_ACCENT = (16, 185, 129)        # #10b981 – emerald-500
_DANGER = (239, 68, 68)         # #ef4444 – red-500
_INDIGO = (99, 102, 241)        # #6366f1 – indigo-500
_WHITE = (255, 255, 255)
_GRAY_50 = (248, 250, 252)      # #f8fafc
_GRAY_100 = (241, 245, 249)     # #f1f5f9
_GRAY_700 = (51, 65, 85)        # #334155
_BLACK = (0, 0, 0)


# ===================================================================
# PDF
# ===================================================================

class _ReportPDF(FPDF):
    """FPDF subclass com header/footer personalizados."""

    def __init__(
        self,
        period_from: date | None,
        period_to: date | None,
    ) -> None:
        super().__init__(orientation="P", unit="mm", format="A4")
        self.period_from = period_from
        self.period_to = period_to
        self.set_auto_page_break(auto=True, margin=20)

    # ---- Header -------------------------------------------------------
    def header(self) -> None:
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(*_PRIMARY_DARK)
        self.cell(0, 10, "ITES - Núcleo de Práticas Jurídicas", align="C", new_x="LMARGIN", new_y="NEXT")

        self.set_font("Helvetica", "", 11)
        self.set_text_color(*_GRAY_700)
        self.cell(0, 6, "Relatório do Núcleo Jurídico", align="C", new_x="LMARGIN", new_y="NEXT")

        if self.period_from or self.period_to:
            de = self.period_from.strftime("%d/%m/%Y") if self.period_from else "início"
            ate = self.period_to.strftime("%d/%m/%Y") if self.period_to else "hoje"
            self.set_font("Helvetica", "I", 9)
            self.cell(0, 5, f"Período: {de}  a  {ate}", align="C", new_x="LMARGIN", new_y="NEXT")

        # Linha decorativa
        self.set_draw_color(*_PRIMARY)
        self.set_line_width(0.6)
        y = self.get_y() + 2
        self.line(10, y, 200, y)
        self.set_y(y + 4)

    # ---- Footer -------------------------------------------------------
    def footer(self) -> None:
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(160, 160, 160)
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        self.cell(0, 10, f"Gerado em {now}", align="L")
        self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", align="R")

    # ---- Helpers ------------------------------------------------------
    def section_title(self, title: str) -> None:
        """Título de seção com fundo colorido."""
        self.ln(4)
        self.set_fill_color(*_PRIMARY)
        self.set_text_color(*_WHITE)
        self.set_font("Helvetica", "B", 11)
        self.cell(0, 8, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def _table_header(self, widths: list[float], labels: list[str]) -> None:
        self.set_fill_color(*_PRIMARY_DARK)
        self.set_text_color(*_WHITE)
        self.set_font("Helvetica", "B", 9)
        for w, lbl in zip(widths, labels):
            self.cell(w, 7, lbl, border=1, fill=True, align="C")
        self.ln()

    def _table_row(
        self,
        widths: list[float],
        values: list[str],
        idx: int,
        aligns: list[str] | None = None,
    ) -> None:
        if aligns is None:
            aligns = ["L"] + ["C"] * (len(values) - 1)
        if idx % 2 == 0:
            self.set_fill_color(*_WHITE)
        else:
            self.set_fill_color(*_GRAY_50)
        self.set_text_color(*_BLACK)
        self.set_font("Helvetica", "", 9)
        for w, val, al in zip(widths, values, aligns):
            self.cell(w, 6, val, border=1, fill=True, align=al)
        self.ln()


def generate_pdf(
    summary: dict[str, Any],
    by_status: list[dict[str, Any]],
    by_area: list[dict[str, Any]],
    by_student: list[dict[str, Any]],
    by_teacher: list[dict[str, Any]],
    period_from: date | None,
    period_to: date | None,
) -> bytes:
    """Gera relatório completo em PDF e retorna os bytes do arquivo."""

    pdf = _ReportPDF(period_from, period_to)
    pdf.alias_nb_pages()
    pdf.add_page()

    # ------------------------------------------------------------------
    # 1. Resumo Geral
    # ------------------------------------------------------------------
    pdf.section_title("1. Resumo Geral")

    counters = summary.get("counters", {})
    finalizados = counters.get("finalizado", 0) + counters.get("arquivado", 0)
    em_analise = (
        counters.get("encaminhado_ao_professor", 0)
        + counters.get("em_analise_pelo_professor", 0)
        + counters.get("correcao_solicitada", 0)
    )

    metrics = [
        ("Total de Atendimentos", str(summary.get("total", 0)), _PRIMARY),
        ("Urgentes", str(summary.get("urgentes", 0)), _DANGER),
        ("Finalizados", str(finalizados), _ACCENT),
        ("Em Análise", str(em_analise), _INDIGO),
    ]

    card_w = 45
    card_gap = 2.5
    start_x = (210 - (card_w * 4 + card_gap * 3)) / 2
    y0 = pdf.get_y()

    for i, (label, value, color) in enumerate(metrics):
        x = start_x + i * (card_w + card_gap)
        pdf.set_xy(x, y0)
        pdf.set_fill_color(*color)
        pdf.set_text_color(*_WHITE)
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(card_w, 14, value, fill=True, align="C", new_x="LMARGIN", new_y="TOP")
        pdf.set_xy(x, y0 + 14)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(card_w, 6, label, fill=True, align="C")

    pdf.set_y(y0 + 24)

    # ------------------------------------------------------------------
    # 2. Por Status
    # ------------------------------------------------------------------
    pdf.section_title("2. Por Status")
    widths_status = [120.0, 70.0]
    pdf._table_header(widths_status, ["Status", "Quantidade"])
    for idx, row in enumerate(by_status):
        pdf._table_row(
            widths_status,
            [row.get("label", row.get("status", "")), str(row.get("count", 0))],
            idx,
        )

    # ------------------------------------------------------------------
    # 3. Por Área Jurídica
    # ------------------------------------------------------------------
    pdf.section_title("3. Por Área Jurídica")
    widths_area = [120.0, 70.0]
    pdf._table_header(widths_area, ["Área Jurídica", "Quantidade"])
    for idx, row in enumerate(by_area):
        pdf._table_row(
            widths_area,
            [row.get("legal_area_name") or "Não informado", str(row.get("count", 0))],
            idx,
        )

    # ------------------------------------------------------------------
    # 4. Produtividade por Aluno
    # ------------------------------------------------------------------
    pdf.section_title("4. Produtividade por Aluno")
    widths_user = [66.0, 31.0, 31.0, 31.0, 31.0]
    user_headers = ["Nome", "Total", "Em andamento", "Finalizados", "Urgentes"]
    pdf._table_header(widths_user, user_headers)
    for idx, row in enumerate(by_student):
        pdf._table_row(
            widths_user,
            [
                row.get("user_name", ""),
                str(row.get("total", 0)),
                str(row.get("em_andamento", 0)),
                str(row.get("finalizados", 0)),
                str(row.get("urgentes", 0)),
            ],
            idx,
        )

    # ------------------------------------------------------------------
    # 5. Produtividade por Professor
    # ------------------------------------------------------------------
    pdf.section_title("5. Produtividade por Professor")
    pdf._table_header(widths_user, user_headers)
    for idx, row in enumerate(by_teacher):
        pdf._table_row(
            widths_user,
            [
                row.get("user_name", ""),
                str(row.get("total", 0)),
                str(row.get("em_andamento", 0)),
                str(row.get("finalizados", 0)),
                str(row.get("urgentes", 0)),
            ],
            idx,
        )

    return pdf.output()


# ===================================================================
# Excel
# ===================================================================

# Estilos reutilizáveis
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2563EB")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws: Any, row: int, col_count: int) -> None:
    """Aplica estilo à linha de cabeçalho."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _style_data_cell(cell: Any, align: str = "center") -> None:
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER if align == "center" else _LEFT


def _write_title(ws: Any, title: str, col_span: int) -> None:
    """Escreve título na primeira linha e mescla colunas."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws: Any, col_count: int, min_width: int = 12) -> None:
    """Ajusta largura das colunas com base no conteúdo."""
    for col in range(1, col_count + 1):
        max_len = min_width
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, 50)


def generate_excel(
    summary: dict[str, Any],
    by_status: list[dict[str, Any]],
    by_area: list[dict[str, Any]],
    by_student: list[dict[str, Any]],
    by_teacher: list[dict[str, Any]],
    period_from: date | None,
    period_to: date | None,
) -> bytes:
    """Gera relatório completo em Excel e retorna os bytes do arquivo."""

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 – Resumo
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Resumo"

    _write_title(ws, "Resumo Geral", 2)

    counters = summary.get("counters", {})
    finalizados = counters.get("finalizado", 0) + counters.get("arquivado", 0)
    em_analise = (
        counters.get("encaminhado_ao_professor", 0)
        + counters.get("em_analise_pelo_professor", 0)
        + counters.get("correcao_solicitada", 0)
    )

    header_row = 3
    headers = ["Métrica", "Valor"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=ci, value=h)
    _style_header_row(ws, header_row, len(headers))

    data_rows = [
        ("Total de Atendimentos", summary.get("total", 0)),
        ("Urgentes", summary.get("urgentes", 0)),
        ("Finalizados", finalizados),
        ("Em Análise", em_analise),
    ]

    if period_from or period_to:
        de = period_from.strftime("%d/%m/%Y") if period_from else "início"
        ate = period_to.strftime("%d/%m/%Y") if period_to else "hoje"
        data_rows.insert(0, ("Período", f"{de} a {ate}"))

    for ri, (metric, value) in enumerate(data_rows, start=header_row + 1):
        c1 = ws.cell(row=ri, column=1, value=metric)
        c2 = ws.cell(row=ri, column=2, value=value)
        _style_data_cell(c1, "left")
        _style_data_cell(c2, "center")

    _auto_width(ws, 2)

    # ------------------------------------------------------------------
    # Sheet 2 – Por Status
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Por Status")
    _write_title(ws2, "Atendimentos por Status", 2)

    headers2 = ["Status", "Quantidade"]
    for ci, h in enumerate(headers2, start=1):
        ws2.cell(row=3, column=ci, value=h)
    _style_header_row(ws2, 3, len(headers2))

    for ri, row in enumerate(by_status, start=4):
        c1 = ws2.cell(row=ri, column=1, value=row.get("label", row.get("status", "")))
        c2 = ws2.cell(row=ri, column=2, value=row.get("count", 0))
        _style_data_cell(c1, "left")
        _style_data_cell(c2, "center")

    _auto_width(ws2, 2, min_width=20)

    # ------------------------------------------------------------------
    # Sheet 3 – Por Área
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Por Área")
    _write_title(ws3, "Atendimentos por Área Jurídica", 2)

    headers3 = ["Área Jurídica", "Quantidade"]
    for ci, h in enumerate(headers3, start=1):
        ws3.cell(row=3, column=ci, value=h)
    _style_header_row(ws3, 3, len(headers3))

    for ri, row in enumerate(by_area, start=4):
        c1 = ws3.cell(row=ri, column=1, value=row.get("legal_area_name") or "Não informado")
        c2 = ws3.cell(row=ri, column=2, value=row.get("count", 0))
        _style_data_cell(c1, "left")
        _style_data_cell(c2, "center")

    _auto_width(ws3, 2, min_width=20)

    # ------------------------------------------------------------------
    # Sheet 4 – Produtividade Alunos
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("Produtividade Alunos")
    _write_title(ws4, "Produtividade por Aluno", 5)

    headers4 = ["Nome", "Total", "Em andamento", "Finalizados", "Urgentes"]
    for ci, h in enumerate(headers4, start=1):
        ws4.cell(row=3, column=ci, value=h)
    _style_header_row(ws4, 3, len(headers4))

    for ri, row in enumerate(by_student, start=4):
        vals = [
            row.get("user_name", ""),
            row.get("total", 0),
            row.get("em_andamento", 0),
            row.get("finalizados", 0),
            row.get("urgentes", 0),
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws4.cell(row=ri, column=ci, value=v)
            _style_data_cell(c, "left" if ci == 1 else "center")

    _auto_width(ws4, 5)

    # ------------------------------------------------------------------
    # Sheet 5 – Produtividade Professores
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("Produtividade Professores")
    _write_title(ws5, "Produtividade por Professor", 5)

    headers5 = ["Nome", "Total", "Em andamento", "Finalizados", "Urgentes"]
    for ci, h in enumerate(headers5, start=1):
        ws5.cell(row=3, column=ci, value=h)
    _style_header_row(ws5, 3, len(headers5))

    for ri, row in enumerate(by_teacher, start=4):
        vals = [
            row.get("user_name", ""),
            row.get("total", 0),
            row.get("em_andamento", 0),
            row.get("finalizados", 0),
            row.get("urgentes", 0),
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws5.cell(row=ri, column=ci, value=v)
            _style_data_cell(c, "left" if ci == 1 else "center")

    _auto_width(ws5, 5)

    # ------------------------------------------------------------------
    # Serializar para bytes
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
